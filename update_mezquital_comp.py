import pandas as pd
import openpyxl

# 1. Get Data from CSV
csv_path = r'c:\Descargas_SRP\SRP-SR-2025_14-05-2026 07-32-54.csv'
df = pd.read_csv(csv_path)
df['MUNICIPIO'] = df['MUNICIPIO'].astype(str).str.strip().str.upper()
mez = df[df['MUNICIPIO'] == 'MEZQUITAL'].copy()

# Ensure numeric
numeric_cols = ['SRP  PRIMERA TOTAL', 'SRP SEGUNDA TOTAL', 'SR PRIMERA TOTAL', 'SR SEGUNDA TOTAL']
for col in numeric_cols:
    mez[col] = pd.to_numeric(mez[col], errors='coerce').fillna(0)

def get_stats(season, inst):
    subset = mez[(mez['Temporada'] == season) & (mez['INSTITUCION'] == inst)]
    if subset.empty:
        return [0, 0, 0, 0]
    return [
        subset['SRP  PRIMERA TOTAL'].sum(),
        subset['SRP SEGUNDA TOTAL'].sum(),
        subset['SR PRIMERA TOTAL'].sum(),
        subset['SR SEGUNDA TOTAL'].sum()
    ]

# 2. Update Excel
excel_path = 'Mezquital_Comp_Temp.xlsx'
wb = openpyxl.load_workbook(excel_path)
sheet = wb['Comparativo 2025 vs 2026']

# Mapping: (Season, Inst) -> Starting Row
mapping = {
    (2025, 'IMSS B'): 6,
    (2025, 'SSA'): 7,
    (2026, 'IMSS B'): 10,
    (2026, 'SSA'): 11
}

for (season, inst), start_row in mapping.items():
    vals = get_stats(season, inst)
    # SRP 1, SRP 2, SR 1, SR 2
    sheet.cell(row=start_row, column=2).value = int(vals[0])
    sheet.cell(row=start_row, column=3).value = int(vals[1])
    sheet.cell(row=start_row, column=4).value = int(vals[2])
    sheet.cell(row=start_row, column=5).value = int(vals[3])
    
    # Totals
    total_srp = vals[0] + vals[1]
    total_sr = vals[2] + vals[3]
    gran_total = total_srp + total_sr
    
    sheet.cell(row=start_row, column=6).value = int(total_srp)
    sheet.cell(row=start_row, column=7).value = int(total_sr)
    sheet.cell(row=start_row, column=8).value = int(gran_total)

# Recalculate Season Totals (Row 8 and 12)
def update_season_total(row_num, source_rows):
    for col in range(2, 9):
        sum_val = sum([sheet.cell(row=r, column=col).value or 0 for r in source_rows])
        sheet.cell(row=row_num, column=col).value = int(sum_val)

update_season_total(8, [6, 7])
update_season_total(12, [10, 11])

wb.save('Mezquital_Comparativo_Actualizado.xlsx')
print("Successfully updated Mezquital comparative data.")
