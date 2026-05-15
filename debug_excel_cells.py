import openpyxl

wb = openpyxl.load_workbook('Mezquital_Comp_Temp.xlsx')
sheet = wb['Comparativo 2025 vs 2026']

print("Cell Contents (Rows 1-15, Cols 1-2):")
for r in range(1, 16):
    val = sheet.cell(row=r, column=1).value
    type_info = type(sheet.cell(row=r, column=1)).__name__
    print(f"Row {r}, Col 1: '{val}' ({type_info})")
    
    val2 = sheet.cell(row=r, column=2).value
    type_info2 = type(sheet.cell(row=r, column=2)).__name__
    print(f"Row {r}, Col 2: '{val2}' ({type_info2})")
