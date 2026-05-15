import openpyxl

excel_path = 'Mezquital_Comp_Fixing.xlsx'
wb = openpyxl.load_workbook(excel_path)
sheet = wb['Comparativo 2025 vs 2026']

# Row 8: TOTAL TEMPORADA 2025
# Row 12: TOTAL TEMPORADA 2026
# Row 13: VARIACIÓN (2026 - 2025)
# Row 14: TOTAL AMBAS TEMPORADA

for col in range(2, 9):
    v2025 = sheet.cell(row=8, column=col).value or 0
    v2026 = sheet.cell(row=12, column=col).value or 0
    
    # Update Variation (Row 13)
    variation = v2026 - v2025
    sheet.cell(row=13, column=col).value = variation
    
    # Update Total Combined (Row 14)
    total_combined = v2025 + v2026
    sheet.cell(row=14, column=col).value = int(total_combined)

wb.save('Mezquital_Comparativo_Corregido_Final.xlsx')
print("Successfully fixed the sum formulas and values in the comparative table.")
