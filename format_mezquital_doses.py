import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import re
from datetime import datetime

# 1. Load Data
path = 'Dosis_por_Localidad_Mezquital_Actualizado.xlsx'
df = pd.read_excel(path)

# 2. Fix Percentages (Cap at 1.0)
df['COBERTURA (%)'] = df['COBERTURA (%)'].apply(lambda x: min(1.0, x) if isinstance(x, (int, float)) else x)

# 3. Sort by Date
# Helper to extract a sortable date from the "Fecha(s) de Atención" column
def extract_sort_date(date_str):
    if pd.isna(date_str):
        return "2099-01-01"
    date_str = str(date_str)
    # Try to find a date like YYYY-MM-DD
    match = re.search(r'(\d{4})-(\d{2})-(\d{2})', date_str)
    if match:
        return f"{match.group(1)}-{match.group(2)}-{match.group(3)}"
    
    # Try to find DD/MM/YYYY
    match = re.search(r'(\d{2})/(\d{2})/(\d{4})', date_str)
    if match:
        return f"{match.group(3)}-{match.group(2)}-{match.group(1)}"
    
    # Try to find "ABRIL2026" or similar
    months = {
        'ENERO': '01', 'FEBRERO': '02', 'MARZO': '03', 'ABRIL': '04', 'MAYO': '05', 'JUNIO': '06',
        'JULIO': '07', 'AGOSTO': '08', 'SEPTIEMBRE': '09', 'OCTUBRE': '10', 'NOVIEMBRE': '11', 'DICIEMBRE': '12'
    }
    for m_name, m_num in months.items():
        if m_name in date_str.upper():
            year_match = re.search(r'20\d{2}', date_str)
            year = year_match.group(0) if year_match else "2026"
            return f"{year}-{m_num}-01"
            
    return "2026-12-31"

df['sort_key'] = df['Fecha(s) de Atención'].apply(extract_sort_date)
df = df.sort_values(by='sort_key').drop(columns=['sort_key'])

# 4. Save with Styling
output_path = 'Dosis_Mezquital_Formateado.xlsx'
df.to_excel(output_path, index=False)

# Re-open with openpyxl for styling
wb = openpyxl.load_workbook(output_path)
ws = wb.active

# Styling Constants
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
row_fill_alt = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Format Header
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")
    cell.border = border

# Format Rows
for r_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=2):
    fill = row_fill_alt if r_idx % 2 == 0 else None
    for cell in row:
        if fill:
            cell.fill = fill
        cell.border = border
        
        # Percentage Formatting for last column
        if cell.column == ws.max_column:
            cell.number_format = '0.0%'

# Adjust column widths
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    ws.column_dimensions[column].width = max_length + 2

wb.save(output_path)
print("Formatting and sorting complete.")
